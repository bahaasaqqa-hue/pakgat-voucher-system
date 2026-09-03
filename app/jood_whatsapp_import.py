from __future__ import annotations

import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path

from app.jood_company_ops import CONTACT_TYPES, normalize_contact_phone


@dataclass(frozen=True)
class ImportedContact:
    phone: str
    contact_type: str
    display_name: str = ""
    business_name: str = ""
    city: str = ""
    notes: str = ""


HEADER_ALIASES = {
    "phone": {"phone", "mobile", "رقم", "الجوال", "رقم الجوال", "الهاتف"},
    "display_name": {"name", "display_name", "الاسم", "اسم العميل", "اسم المسؤول"},
    "business_name": {"business", "business_name", "company", "اسم النشاط", "النشاط", "الشركة"},
    "city": {"city", "المدينة"},
    "notes": {"notes", "note", "ملاحظات", "ملاحظة"},
    "contact_type": {"type", "contact_type", "النوع", "نوع العميل"},
}


def _clean(value) -> str:
    return " ".join(str(value or "").strip().split())


def _field_for_header(value: str) -> str:
    clean = _clean(value).lower()
    for field, aliases in HEADER_ALIASES.items():
        if clean in aliases:
            return field
    return ""


def _rows_from_csv(body: bytes) -> list[list[str]]:
    try:
        text = body.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("CSV must be UTF-8 encoded") from exc
    return [list(row) for row in csv.reader(StringIO(text))]


def _rows_from_xlsx(body: bytes) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("XLSX support is not installed") from exc
    try:
        workbook = load_workbook(BytesIO(body), read_only=True, data_only=True)
        sheet = workbook.active
        return [[_clean(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    except Exception as exc:
        raise ValueError("Invalid XLSX file") from exc


def parse_contact_upload(filename: str, body: bytes, default_type: str) -> list[ImportedContact]:
    suffix = Path(str(filename or "")).suffix.lower()
    if suffix == ".csv":
        raw_rows = _rows_from_csv(body)
    elif suffix == ".xlsx":
        raw_rows = _rows_from_xlsx(body)
    else:
        raise ValueError("Upload must be a CSV or XLSX file")

    rows = [list(map(_clean, row)) for row in raw_rows if any(_clean(value) for value in row)]
    if not rows:
        raise ValueError("Upload contains no valid contacts")

    mapped_headers = [_field_for_header(value) for value in rows[0]]
    has_header = "phone" in mapped_headers
    if has_header:
        indexes = {field: index for index, field in enumerate(mapped_headers) if field}
        data_rows = rows[1:]
    else:
        indexes = {"phone": 0, "display_name": 1, "business_name": 2, "city": 3, "notes": 4}
        data_rows = rows

    fallback_type = default_type if default_type in CONTACT_TYPES else "customer"
    contacts: list[ImportedContact] = []
    seen: set[str] = set()
    for row in data_rows:
        value = lambda field: row[indexes[field]] if field in indexes and indexes[field] < len(row) else ""
        phone = normalize_contact_phone(value("phone"))
        if not phone or phone in seen:
            continue
        contact_type = value("contact_type").lower()
        if contact_type not in CONTACT_TYPES:
            contact_type = fallback_type
        seen.add(phone)
        contacts.append(
            ImportedContact(
                phone=phone,
                contact_type=contact_type,
                display_name=value("display_name")[:255],
                business_name=value("business_name")[:255],
                city=value("city")[:120],
                notes=value("notes")[:4000],
            )
        )
    if not contacts:
        raise ValueError("Upload contains no valid contacts")
    return contacts

